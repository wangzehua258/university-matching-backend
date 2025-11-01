#!/usr/bin/env python3
"""
数据验证脚本 - 验证Excel文件中的数据格式是否正确

使用方法：
    python scripts/validate_data.py --country AU
    python scripts/validate_data.py --country UK
    python scripts/validate_data.py --country SG
    python scripts/validate_data.py --all  # 验证所有国家
"""

import os
import sys
from pathlib import Path

# 添加项目根目录到Python路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
except ImportError:
    print("❌ 需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


def validate_au_data(file_path):
    """验证澳大利亚数据"""
    print(f"\n📋 验证澳大利亚数据: {file_path}")
    
    expected_headers = [
        "name", "country", "city", "rank", "tuition_local", "currency", "tuition_usd",
        "study_length_years", "intakes", "english_requirements", "requires_english_test",
        "group_of_eight", "work_integrated_learning", "placement_rate", "post_study_visa_years",
        "scholarship_available", "strengths", "tags", "intlRate", "website"
    ]
    
    errors = []
    warnings = []
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 检查表头
        headers = [cell.value for cell in ws[1]]
        if headers != expected_headers:
            errors.append(f"❌ 表头不匹配！期望 {len(expected_headers)} 列，实际 {len(headers)} 列")
            errors.append(f"   期望: {', '.join(expected_headers)}")
            errors.append(f"   实际: {', '.join([str(h) for h in headers[:10]])}...")
            return errors, warnings
        
        print(f"✅ 表头正确（共 {len(expected_headers)} 列）")
        
        # 验证数据行
        data_rows = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if not any(cell.value for cell in row):
                continue  # 跳过空行
            
            data_rows += 1
            row_values = [cell.value for cell in row]
            
            # 检查必填字段
            if not row_values[0]:  # name
                errors.append(f"❌ 第{row_idx}行: name 不能为空")
            
            # 检查数字字段
            try:
                rank = int(row_values[3]) if row_values[3] else None
                if rank is None or rank <= 0:
                    errors.append(f"❌ 第{row_idx}行: rank 必须是正整数")
            except (ValueError, TypeError):
                errors.append(f"❌ 第{row_idx}行: rank 格式错误 ({row_values[3]})")
            
            try:
                tuition_local = int(row_values[4]) if row_values[4] else None
                if tuition_local is None or tuition_local <= 0:
                    errors.append(f"❌ 第{row_idx}行: tuition_local 必须是正整数")
            except (ValueError, TypeError):
                errors.append(f"❌ 第{row_idx}行: tuition_local 格式错误 ({row_values[4]})")
            
            # 检查布尔字段
            bool_fields = {
                10: "requires_english_test",
                11: "group_of_eight",
                12: "work_integrated_learning",
                15: "scholarship_available"
            }
            for col_idx, field_name in bool_fields.items():
                value = row_values[col_idx]
                if value not in [True, False, "TRUE", "FALSE", "true", "false", 1, 0]:
                    errors.append(f"❌ 第{row_idx}行: {field_name} 必须是 TRUE/FALSE，当前值: {value}")
            
            # 检查浮点数字段
            try:
                intl_rate = float(row_values[18]) if row_values[18] is not None else None
                if intl_rate is not None and (intl_rate < 0 or intl_rate > 1):
                    warnings.append(f"⚠️  第{row_idx}行: intlRate 应该在 0-1 之间，当前值: {intl_rate}")
            except (ValueError, TypeError):
                if row_values[18] is not None:
                    errors.append(f"❌ 第{row_idx}行: intlRate 格式错误 ({row_values[18]})")
        
        print(f"✅ 检查了 {data_rows} 行数据")
        
    except FileNotFoundError:
        errors.append(f"❌ 文件不存在: {file_path}")
    except Exception as e:
        errors.append(f"❌ 读取文件出错: {e}")
    
    return errors, warnings


def validate_uk_data(file_path):
    """验证英国数据"""
    print(f"\n📋 验证英国数据: {file_path}")
    
    expected_headers = [
        "name", "country", "city", "rank", "tuition_local", "currency", "tuition_usd",
        "study_length_years", "ucas_deadline_type", "typical_offer_alevel", "typical_offer_ib",
        "foundation_available", "russell_group", "placement_year_available", "interview_required",
        "admissions_tests", "personal_statement_weight", "strengths", "tags", "intlRate",
        "website", "scholarship_available"
    ]
    
    errors = []
    warnings = []
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        if headers != expected_headers:
            errors.append(f"❌ 表头不匹配！期望 {len(expected_headers)} 列，实际 {len(headers)} 列")
            return errors, warnings
        
        print(f"✅ 表头正确（共 {len(expected_headers)} 列）")
        
        data_rows = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if not any(cell.value for cell in row):
                continue
            
            data_rows += 1
            row_values = [cell.value for cell in row]
            
            if not row_values[0]:
                errors.append(f"❌ 第{row_idx}行: name 不能为空")
            
            # 检查布尔字段
            bool_fields = {
                11: "foundation_available",
                12: "russell_group",
                13: "placement_year_available",
                14: "interview_required",
                21: "scholarship_available"
            }
            for col_idx, field_name in bool_fields.items():
                value = row_values[col_idx]
                if value not in [True, False, "TRUE", "FALSE", "true", "false", 1, 0]:
                    errors.append(f"❌ 第{row_idx}行: {field_name} 必须是 TRUE/FALSE，当前值: {value}")
            
            # 检查 personal_statement_weight
            try:
                ps_weight = int(row_values[16]) if row_values[16] is not None else None
                if ps_weight is None or ps_weight < 1 or ps_weight > 10:
                    errors.append(f"❌ 第{row_idx}行: personal_statement_weight 必须是 1-10 的整数，当前值: {ps_weight}")
            except (ValueError, TypeError):
                if row_values[16] is not None:
                    errors.append(f"❌ 第{row_idx}行: personal_statement_weight 格式错误 ({row_values[16]})")
        
        print(f"✅ 检查了 {data_rows} 行数据")
        
    except FileNotFoundError:
        errors.append(f"❌ 文件不存在: {file_path}")
    except Exception as e:
        errors.append(f"❌ 读取文件出错: {e}")
    
    return errors, warnings


def validate_sg_data(file_path):
    """验证新加坡数据"""
    print(f"\n📋 验证新加坡数据: {file_path}")
    
    expected_headers = [
        "name", "country", "city", "rank", "tuition_local", "currency", "tuition_usd",
        "study_length_years", "tuition_grant_available", "tuition_grant_bond_years",
        "interview_required", "essay_or_portfolio_required", "coop_or_internship_required",
        "industry_links_score", "exchange_opportunities_score", "strengths", "tags",
        "intlRate", "website", "scholarship_available"
    ]
    
    errors = []
    warnings = []
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        if headers != expected_headers:
            errors.append(f"❌ 表头不匹配！期望 {len(expected_headers)} 列，实际 {len(headers)} 列")
            return errors, warnings
        
        print(f"✅ 表头正确（共 {len(expected_headers)} 列）")
        
        data_rows = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if not any(cell.value for cell in row):
                continue
            
            data_rows += 1
            row_values = [cell.value for cell in row]
            
            if not row_values[0]:
                errors.append(f"❌ 第{row_idx}行: name 不能为空")
            
            # 检查布尔字段
            bool_fields = {
                8: "tuition_grant_available",
                10: "interview_required",
                11: "essay_or_portfolio_required",
                12: "coop_or_internship_required",
                19: "scholarship_available"
            }
            for col_idx, field_name in bool_fields.items():
                value = row_values[col_idx]
                if value not in [True, False, "TRUE", "FALSE", "true", "false", 1, 0]:
                    errors.append(f"❌ 第{row_idx}行: {field_name} 必须是 TRUE/FALSE，当前值: {value}")
            
            # 检查 industry_links_score
            try:
                score = int(row_values[13]) if row_values[13] is not None else None
                if score is None or score < 1 or score > 10:
                    errors.append(f"❌ 第{row_idx}行: industry_links_score 必须是 1-10 的整数，当前值: {score}")
            except (ValueError, TypeError):
                if row_values[13] is not None:
                    errors.append(f"❌ 第{row_idx}行: industry_links_score 格式错误 ({row_values[13]})")
            
            # 检查 exchange_opportunities_score（可选）
            if row_values[14] is not None:
                try:
                    ex_score = int(row_values[14])
                    if ex_score < 1 or ex_score > 10:
                        warnings.append(f"⚠️  第{row_idx}行: exchange_opportunities_score 建议在 1-10 之间，当前值: {ex_score}")
                except (ValueError, TypeError):
                    warnings.append(f"⚠️  第{row_idx}行: exchange_opportunities_score 格式可能有问题 ({row_values[14]})")
        
        print(f"✅ 检查了 {data_rows} 行数据")
        
    except FileNotFoundError:
        errors.append(f"❌ 文件不存在: {file_path}")
    except Exception as e:
        errors.append(f"❌ 读取文件出错: {e}")
    
    return errors, warnings


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="验证国际大学数据Excel文件")
    parser.add_argument("--country", choices=["AU", "UK", "SG"], help="指定国家代码")
    parser.add_argument("--all", action="store_true", help="验证所有国家")
    
    args = parser.parse_args()
    
    base_dir = Path(__file__).parent.parent / "data" / "international"
    
    results = {}
    
    if args.all or args.country == "AU":
        au_file = base_dir / "AUSTRALIA.xlsx"
        errors, warnings = validate_au_data(au_file)
        results["Australia"] = (errors, warnings)
    
    if args.all or args.country == "UK":
        uk_file = base_dir / "UK.xlsx"
        errors, warnings = validate_uk_data(uk_file)
        results["United Kingdom"] = (errors, warnings)
    
    if args.all or args.country == "SG":
        sg_file = base_dir / "SINGAPORE.xlsx"
        errors, warnings = validate_sg_data(sg_file)
        results["Singapore"] = (errors, warnings)
    
    # 汇总结果
    print("\n" + "="*60)
    print("📊 验证结果汇总")
    print("="*60)
    
    total_errors = 0
    total_warnings = 0
    
    for country, (errors, warnings) in results.items():
        print(f"\n{country}:")
        if errors:
            print(f"  ❌ 错误: {len(errors)} 个")
            for err in errors[:5]:  # 只显示前5个错误
                print(f"    {err}")
            if len(errors) > 5:
                print(f"    ... 还有 {len(errors) - 5} 个错误")
            total_errors += len(errors)
        else:
            print("  ✅ 无错误")
        
        if warnings:
            print(f"  ⚠️  警告: {len(warnings)} 个")
            for warn in warnings[:3]:  # 只显示前3个警告
                print(f"    {warn}")
            if len(warnings) > 3:
                print(f"    ... 还有 {len(warnings) - 3} 个警告")
            total_warnings += len(warnings)
        else:
            print("  ✅ 无警告")
    
    print("\n" + "="*60)
    if total_errors == 0 and total_warnings == 0:
        print("✅ 所有数据验证通过！可以导入数据库。")
    elif total_errors == 0:
        print(f"⚠️  有 {total_warnings} 个警告，建议检查。")
    else:
        print(f"❌ 发现 {total_errors} 个错误，{total_warnings} 个警告。请修正后重试。")
    print("="*60)


if __name__ == "__main__":
    main()

