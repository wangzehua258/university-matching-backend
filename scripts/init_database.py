#!/usr/bin/env python3
"""
数据库初始化脚本
支持CSV导入、增量更新和批量操作
"""

import os
import sys
import json
import csv
from pathlib import Path
from pymongo import MongoClient, ASCENDING, DESCENDING
from dotenv import load_dotenv

# 添加项目根目录到Python路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

load_dotenv()

def connect_database():
    """连接MongoDB数据库"""
    mongo_url = os.getenv("MONGO_URL", "mongodb://localhost:27017")
    client = MongoClient(mongo_url)
    db = client.university_matcher
    return db

def create_indexes(db):
    """创建数据库索引"""
    print("创建数据库索引...")
    
    try:
        # 用户集合索引
        db.users.create_index("created_at")
        print("✅ 用户索引创建完成")
    except Exception as e:
        print(f"⚠️  用户索引创建跳过: {e}")
    
    try:
        # 大学集合索引 - 更新以支持新字段
        # 先删除可能冲突的索引
        try:
            db.universities.drop_index("name_1")
            print("🔄 删除旧名称索引")
        except:
            pass
        
        # 创建新索引
        db.universities.create_index("name", unique=True)  # 确保大学名称唯一
        db.universities.create_index("country")
        db.universities.create_index("rank")
        db.universities.create_index([("country", ASCENDING), ("rank", ASCENDING)])
        db.universities.create_index("strengths")
        db.universities.create_index("tuition")
        db.universities.create_index("type")
        db.universities.create_index("school_size")
        db.universities.create_index("tags")
    
        # 新增字段索引
        db.universities.create_index("supports_ed")
        db.universities.create_index("supports_ea")
        db.universities.create_index("supports_rd")
        db.universities.create_index("internship_support_score")
        db.universities.create_index("acceptanceRate")
        db.universities.create_index("intlRate")
        db.universities.create_index("state")
        db.universities.create_index("personality_types")
        
        print("✅ 大学索引创建完成")
    except Exception as e:
        print(f"⚠️  大学索引创建跳过: {e}")
    
    try:
        # 评估结果索引
        db.parent_evaluations.create_index("user_id")
        db.parent_evaluations.create_index("created_at")
        db.student_personality_tests.create_index("user_id")
        db.student_personality_tests.create_index("created_at")
        print("✅ 评估索引创建完成")
    except Exception as e:
        print(f"⚠️  评估索引创建跳过: {e}")
    
    print("索引创建完成")

    # 国际大学集合索引
    try:
        # AU
        db.university_au.create_index("name", unique=True)
        db.university_au.create_index("city")
        db.university_au.create_index("rank")
        db.university_au.create_index("work_integrated_learning")
        db.university_au.create_index("group_of_eight")
        db.university_au.create_index("strengths")
        db.university_au.create_index("tags")
        # UK
        db.university_uk.create_index("name", unique=True)
        db.university_uk.create_index("city")
        db.university_uk.create_index("rank")
        db.university_uk.create_index("foundation_available")
        db.university_uk.create_index("placement_year_available")
        db.university_uk.create_index("russell_group")
        db.university_uk.create_index("strengths")
        db.university_uk.create_index("tags")
        # SG
        db.university_sg.create_index("name", unique=True)
        db.university_sg.create_index("rank")
        db.university_sg.create_index("tuition_grant_available")
        db.university_sg.create_index("strengths")
        db.university_sg.create_index("tags")
        print("✅ 国际大学索引创建完成")
    except Exception as e:
        print(f"⚠️  国际大学索引创建跳过: {e}")

def clean_boolean_value(value):
    """清理布尔值"""
    if isinstance(value, str):
        value = value.strip().upper()
        if value in ['TRUE', 'T', 'YES', 'Y', '1']:
            return True
        elif value in ['FALSE', 'F', 'NO', 'N', '0']:
            return False
    return False

def clean_numeric_value(value, default=0, is_float=False):
    """清理数值"""
    if not value or value == '':
        return default
    
    try:
        if is_float:
            return float(value)
        else:
            return int(value)
    except (ValueError, TypeError):
        return default

def import_universities_from_csv(db, csv_file_path, clear_existing=False):
    """从CSV文件导入大学数据"""
    if not os.path.exists(csv_file_path):
        print(f"CSV文件不存在: {csv_file_path}")
        return
    
    print(f"从CSV文件导入大学数据: {csv_file_path}")
    
    # 是否清空现有数据
    if clear_existing:
        db.universities.delete_many({})
        print("已清空现有数据")
    
    with open(csv_file_path, 'r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        universities = []
        updated_count = 0
        inserted_count = 0
        
        for row_num, row in enumerate(reader, 1):
            try:
                # 调试：显示关键字段的原始值
                if row_num <= 3:  # 只显示前3行
                    print(f"🔍 第{row_num}行调试信息:")
                    print(f"   acceptanceRate: '{row.get('acceptanceRate', 'NOT_FOUND')}'")
                    print(f"   satRange: '{row.get('satRange', 'NOT_FOUND')}'")
                    print(f"   actRange: '{row.get('actRange', 'NOT_FOUND')}'")
                    print(f"   gpaRange: '{row.get('gpaRange', 'NOT_FOUND')}'")
                    print(f"   applicationDeadline: '{row.get('applicationDeadline', 'NOT_FOUND')}'")
                    print(f"   supports_ed: '{row.get('supports_ed', 'NOT_FOUND')}'")
                    print(f"   supports_ea: '{row.get('supports_ea', 'NOT_FOUND')}'")
                    print(f"   supports_rd: '{row.get('supports_rd', 'NOT_FOUND')}'")
                    print(f"   has_internship_program: '{row.get('has_internship_program', 'NOT_FOUND')}'")
                    print(f"   has_research_program: '{row.get('has_research_program', 'NOT_FOUND')}'")
                    print(f"   internship_support_score: '{row.get('internship_support_score', 'NOT_FOUND')}'")
                    print(f"   schoolSize: '{row.get('schoolSize', 'NOT_FOUND')}'")
                    print(f"   website: '{row.get('website', 'NOT_FOUND')}'")
                    print("   ---")
                
                # 数据清洗和转换 - 适配schools.csv格式
                university = {
                    "name": row.get("name", "").strip(),
                    "country": row.get("country", "").strip(),
                    "state": row.get("state", "").strip(),
                    "rank": clean_numeric_value(row.get("rank"), 999),
                    "tuition": clean_numeric_value(row.get("tuition"), 0),
                    "intlRate": clean_numeric_value(row.get("intlRate"), 0, True),
                    "type": row.get("type", "private").strip(),
                    "schoolSize": row.get("schoolSize", "medium").strip(),
                    "strengths": [s.strip() for s in row.get("strengths", "").split(",") if s.strip()] if row.get("strengths") else [],
                    "gptSummary": row.get("gptSummary", "").strip(),
                    "logoUrl": "",  # 暂时留空，后续可以添加
                    "acceptanceRate": clean_numeric_value(row.get("acceptanceRate"), 0, True),
                    "satRange": row.get("satRange", "").strip(),
                    "actRange": row.get("actRange", "").strip(),
                    "gpaRange": row.get("gpaRange", "").strip(),
                    "applicationDeadline": row.get("applicationDeadline", "").strip(),
                    "website": row.get("website", "").strip(),
                    "supports_ed": clean_boolean_value(row.get("supports_ed")),
                    "supports_ea": clean_boolean_value(row.get("supports_ea")),
                    "supports_rd": clean_boolean_value(row.get("supports_rd")),
                    "has_internship_program": clean_boolean_value(row.get("has_internship_program")),
                    "has_research_program": clean_boolean_value(row.get("has_research_program")),
                    "internship_support_score": clean_numeric_value(row.get("internship_support_score"), 5),
                    "personality_types": [s.strip() for s in row.get("personality_types", "").split(",") if s.strip()] if row.get("personality_types") else [],
                    "tags": [s.strip() for s in row.get("tags", "").split(",") if s.strip()] if row.get("tags") else []
                }
                
                # 调试：显示清洗后的关键字段值
                if row_num <= 3:  # 只显示前3行
                    print(f"🔧 第{row_num}行清洗后数据:")
                    print(f"   acceptanceRate: {university['acceptanceRate']}")
                    print(f"   satRange: '{university['satRange']}'")
                    print(f"   actRange: '{university['actRange']}'")
                    print(f"   gpaRange: '{university['gpaRange']}'")
                    print(f"   applicationDeadline: '{university['applicationDeadline']}'")
                    print(f"   supports_ed: {university['supports_ed']}")
                    print(f"   supports_ea: {university['supports_ea']}")
                    print(f"   supports_rd: {university['supports_rd']}")
                    print(f"   has_internship_program: {university['has_internship_program']}")
                    print(f"   has_research_program: {university['has_research_program']}")
                    print(f"   internship_support_score: {university['internship_support_score']}")
                    print(f"   schoolSize: '{university['schoolSize']}'")
                    print(f"   website: '{university['website']}'")
                    print("   ---")
                
                # 验证必需字段
                if not university["name"]:
                    print(f"⚠️  第{row_num}行：缺少大学名称，跳过")
                    continue
                
                # 检查是否已存在（按名称）
                existing = db.universities.find_one({"name": university["name"]})
                if existing:
                    if clear_existing:
                        # 如果清空模式，直接插入
                        universities.append(university)
                    else:
                        # 更新模式，更新现有记录
                        db.universities.update_one(
                            {"name": university["name"]}, 
                            {"$set": university}
                        )
                        updated_count += 1
                else:
                    universities.append(university)
                
            except Exception as e:
                print(f"❌ 第{row_num}行数据错误: {e}")
                print(f"   行数据: {row}")
                continue
        
        # 批量插入新数据
        if universities:
            try:
                result = db.universities.insert_many(universities)
                inserted_count = len(result.inserted_ids)
                print(f"✅ 成功插入 {inserted_count} 所新大学")
            except Exception as e:
                print(f"❌ 批量插入失败: {e}")
                # 尝试逐个插入
                for uni in universities:
                    try:
                        db.universities.insert_one(uni)
                        inserted_count += 1
                    except Exception as e2:
                        print(f"❌ 插入失败 {uni['name']}: {e2}")
        
        print(f"📊 导入完成：新增 {inserted_count} 所，更新 {updated_count} 所")


def _read_xlsx_rows(file_path, expected_headers):
    try:
        import openpyxl
    except Exception as e:
        raise RuntimeError("需要安装 openpyxl 以读取Excel文件: pip install openpyxl") from e
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))[0:len(expected_headers)]]
    if [h.strip() for h in headers] != expected_headers:
        raise RuntimeError(f"Excel表头不匹配，期望: {expected_headers}，实际: {headers}")
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row[:len(expected_headers)]]
        yield dict(zip(expected_headers, values))


def import_au_from_excel(db, file_path, clear_existing=False):
    expected = [
        "name","country","city","rank","tuition_local","currency","tuition_usd","study_length_years",
        "intakes","english_requirements","requires_english_test","group_of_eight","work_integrated_learning",
        "placement_rate","post_study_visa_years","scholarship_available","strengths","tags","intlRate","website"
    ]
    if clear_existing:
        db.university_au.delete_many({})
    inserted, updated = 0, 0
    for row in _read_xlsx_rows(file_path, expected):
        row["strengths"] = [s.strip() for s in (row.get("strengths") or "").split(",") if s and str(s).strip()]
        row["tags"] = [s.strip() for s in (row.get("tags") or "").split(",") if s and str(s).strip()]
        existing = db.university_au.find_one({"name": row["name"]})
        if existing and not clear_existing:
            db.university_au.update_one({"_id": existing["_id"]}, {"$set": row})
            updated += 1
        else:
            db.university_au.insert_one(row)
            inserted += 1
    print(f"✅ AU 导入完成：新增 {inserted}，更新 {updated}")


def import_uk_from_excel(db, file_path, clear_existing=False):
    expected = [
        "name","country","city","rank","tuition_local","currency","tuition_usd","study_length_years",
        "ucas_deadline_type","typical_offer_alevel","typical_offer_ib","foundation_available","russell_group",
        "placement_year_available","interview_required","admissions_tests","personal_statement_weight","strengths",
        "tags","intlRate","website","scholarship_available"
    ]
    if clear_existing:
        db.university_uk.delete_many({})
    inserted, updated = 0, 0
    for row in _read_xlsx_rows(file_path, expected):
        row["strengths"] = [s.strip() for s in (row.get("strengths") or "").split(",") if s and str(s).strip()]
        row["tags"] = [s.strip() for s in (row.get("tags") or "").split(",") if s and str(s).strip()]
        existing = db.university_uk.find_one({"name": row["name"]})
        if existing and not clear_existing:
            db.university_uk.update_one({"_id": existing["_id"]}, {"$set": row})
            updated += 1
        else:
            db.university_uk.insert_one(row)
            inserted += 1
    print(f"✅ UK 导入完成：新增 {inserted}，更新 {updated}")


def import_sg_from_excel(db, file_path, clear_existing=False):
    expected = [
        "name","country","city","rank","tuition_local","currency","tuition_usd","study_length_years",
        "tuition_grant_available","tuition_grant_bond_years","interview_required","essay_or_portfolio_required",
        "coop_or_internship_required","industry_links_score","exchange_opportunities_score","strengths","tags",
        "intlRate","website","scholarship_available"
    ]
    if clear_existing:
        db.university_sg.delete_many({})
    inserted, updated = 0, 0
    for row in _read_xlsx_rows(file_path, expected):
        row["strengths"] = [s.strip() for s in (row.get("strengths") or "").split(",") if s and str(s).strip()]
        row["tags"] = [s.strip() for s in (row.get("tags") or "").split(",") if s and str(s).strip()]
        existing = db.university_sg.find_one({"name": row["name"]})
        if existing and not clear_existing:
            db.university_sg.update_one({"_id": existing["_id"]}, {"$set": row})
            updated += 1
        else:
            db.university_sg.insert_one(row)
            inserted += 1
    print(f"✅ SG 导入完成：新增 {inserted}，更新 {updated}")

def import_from_json(db, json_file_path, clear_existing=False):
    """从JSON文件导入大学数据"""
    if not os.path.exists(json_file_path):
        print(f"JSON文件不存在: {json_file_path}")
        return
    
    print(f"从JSON文件导入大学数据: {json_file_path}")
    
    try:
        with open(json_file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        if clear_existing:
            db.universities.delete_many({})
            print("已清空现有数据")
        
        if isinstance(data, list):
            universities = data
        elif isinstance(data, dict) and "universities" in data:
            universities = data["universities"]
        else:
            print("❌ JSON格式错误：应该是大学数组或包含universities字段的对象")
            return
        
        # 处理数据
        inserted_count = 0
        updated_count = 0
        
        for uni in universities:
            try:
                # 检查是否已存在
                existing = db.universities.find_one({"name": uni["name"]})
                if existing:
                    if clear_existing:
                        db.universities.insert_one(uni)
                        inserted_count += 1
                    else:
                        db.universities.update_one(
                            {"name": uni["name"]}, 
                            {"$set": uni}
                        )
                        updated_count += 1
                else:
                    db.universities.insert_one(uni)
                    inserted_count += 1
            except Exception as e:
                print(f"❌ 处理大学 {uni.get('name', 'Unknown')} 失败: {e}")
        
        print(f"📊 JSON导入完成：新增 {inserted_count} 所，更新 {updated_count} 所")
        
    except Exception as e:
        print(f"❌ JSON文件读取失败: {e}")

def export_to_csv(db, output_file="universities_export.csv"):
    """导出数据库中的大学数据到CSV文件"""
    print(f"导出大学数据到: {output_file}")
    
    # 创建data目录
    data_dir = Path(__file__).parent.parent / "data"
    data_dir.mkdir(exist_ok=True)
    
    output_path = data_dir / output_file
    
    # 获取所有大学数据
    universities = list(db.universities.find({}, {"_id": 0}).sort("rank", 1))
    
    if not universities:
        print("❌ 数据库中没有大学数据")
        return
    
    # 写入CSV
    with open(output_path, 'w', encoding='utf-8', newline='') as file:
        if universities:
            fieldnames = universities[0].keys()
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(universities)
    
    print(f"✅ 成功导出 {len(universities)} 所大学到 {output_path}")

def show_database_stats(db):
    """显示数据库统计信息"""
    print("\n📊 数据库统计信息:")
    print("-" * 40)
    
    total_universities = db.universities.count_documents({})
    print(f"总大学数量: {total_universities}")
    
    if total_universities > 0:
        # 排名分布
        top10 = db.universities.count_documents({"rank": {"$lte": 10}})
        top20 = db.universities.count_documents({"rank": {"$lte": 20}})
        top50 = db.universities.count_documents({"rank": {"$lte": 50}})
        
        print(f"前10名: {top10} 所")
        print(f"前20名: {top20} 所")
        print(f"前50名: {top50} 所")
        
        # 类型分布
        private_count = db.universities.count_documents({"type": "private"})
        public_count = db.universities.count_documents({"type": "public"})
        print(f"私立大学: {private_count} 所")
        print(f"公立大学: {public_count} 所")
        
        # 规模分布
        small_count = db.universities.count_documents({"schoolSize": "small"})
        medium_count = db.universities.count_documents({"schoolSize": "medium"})
        large_count = db.universities.count_documents({"schoolSize": "large"})
        print(f"小型学校: {small_count} 所")
        print(f"中型学校: {medium_count} 所")
        print(f"大型学校: {large_count} 所")
        
        # 国家分布
        usa_count = db.universities.count_documents({"country": "USA"})
        print(f"美国大学: {usa_count} 所")
        
        # 平均学费
        avg_tuition = db.universities.aggregate([
            {"$group": {"_id": None, "avg": {"$avg": "$tuition"}}}
        ]).next()["avg"]
        print(f"平均学费: ${avg_tuition:,.0f}")

def main():
    """主函数"""
    print("🚀 大学数据库管理工具")
    print("=" * 50)
    
    # 连接数据库
    try:
        db = connect_database()
        print("✅ 数据库连接成功")
    except Exception as e:
        print(f"❌ 数据库连接失败: {e}")
        return
    
    # 创建索引
    create_indexes(db)
    
    # 检查数据文件（美国/默认）
    data_dir = Path(__file__).parent.parent / "data"
    data_dir.mkdir(exist_ok=True)
    
    # 优先检查schools.csv，然后是universities.csv
    schools_csv = data_dir / "schools.csv"
    universities_csv = data_dir / "universities.csv"
    json_file = data_dir / "universities.json"
    
    if schools_csv.exists():
        print(f"📁 找到学校数据文件: {schools_csv}")
        choice = input("是否从schools.csv导入数据？(y/n，默认y): ").strip().lower()
        if choice != 'n':
            clear_choice = input("是否清空现有数据？(y/n，默认n): ").strip().lower()
            clear_existing = clear_choice == 'y'
            import_universities_from_csv(db, str(schools_csv), clear_existing)
    elif universities_csv.exists():
        print(f"📁 找到大学数据文件: {universities_csv}")
        choice = input("是否从universities.csv导入数据？(y/n，默认y): ").strip().lower()
        if choice != 'n':
            clear_choice = input("是否清空现有数据？(y/n，默认n): ").strip().lower()
            clear_existing = clear_choice == 'y'
            import_universities_from_csv(db, str(universities_csv), clear_existing)
    elif json_file.exists():
        print(f"📁 找到JSON文件: {json_file}")
        choice = input("是否从JSON导入数据？(y/n，默认y): ").strip().lower()
        if choice != 'n':
            clear_choice = input("是否清空现有数据？(y/n，默认n): ").strip().lower()
            clear_existing = clear_choice == 'y'
            import_from_json(db, str(json_file), clear_existing)
    # 国际数据（AU/UK/SG） - 优先读取Excel
    intl_dir = data_dir / "international"
    intl_dir.mkdir(exist_ok=True)

    au_xlsx = intl_dir / "AUSTRALIA.xlsx"
    uk_xlsx = intl_dir / "UK.xlsx"
    sg_xlsx = intl_dir / "SINGAPORE.xlsx"

    # 如无Excel，尝试使用已有CSV结构生成Excel样例
    def ensure_sample_excels():
        try:
            import openpyxl
        except Exception:
            print("⚠️ 未安装openpyxl，跳过生成Excel样例。可安装后重试: pip install openpyxl")
        return
        if not au_xlsx.exists():
            wb = openpyxl.Workbook(); ws = wb.active
            ws.append(["name","country","city","rank","tuition_local","currency","tuition_usd","study_length_years","intakes","english_requirements","requires_english_test","group_of_eight","work_integrated_learning","placement_rate","post_study_visa_years","scholarship_available","strengths","tags","intlRate","website"])
            ws.append(["University of Melbourne","Australia","Melbourne",32,48000,"AUD",32000,3.0,"Feb, Jul","IELTS 6.5 no<6.0",True,True,True,0.85,4.0,True,"CS,Engineering,Business","Go8,WIL",0.35,"https://www.unimelb.edu.au"])  # 示例
            wb.save(au_xlsx)
            print(f"🧪 已生成样例: {au_xlsx}")
        if not uk_xlsx.exists():
            wb = openpyxl.Workbook(); ws = wb.active
            ws.append(["name","country","city","rank","tuition_local","currency","tuition_usd","study_length_years","ucas_deadline_type","typical_offer_alevel","typical_offer_ib","foundation_available","russell_group","placement_year_available","interview_required","admissions_tests","personal_statement_weight","strengths","tags","intlRate","website","scholarship_available"])
            ws.append(["University of Manchester","United Kingdom","Manchester",52,28000,"GBP",35000,3.0,"Main(1/31)","AAA","36-38 HL 666",True,True,True,False,"TMUA",7,"CS,Engineering,Business","Russell,Placement",0.30,"https://www.manchester.ac.uk",True])
            wb.save(uk_xlsx)
            print(f"🧪 已生成样例: {uk_xlsx}")
        if not sg_xlsx.exists():
            wb = openpyxl.Workbook(); ws = wb.active
            ws.append(["name","country","city","rank","tuition_local","currency","tuition_usd","study_length_years","tuition_grant_available","tuition_grant_bond_years","interview_required","essay_or_portfolio_required","coop_or_internship_required","industry_links_score","exchange_opportunities_score","strengths","tags","intlRate","website","scholarship_available"])
            ws.append(["National University of Singapore","Singapore","Singapore",11,45000,"SGD",33000,4.0,True,3,True,True,True,9,8,"CS,Engineering,Business","Coop,Exchange",0.28,"https://www.nus.edu.sg",True])
            wb.save(sg_xlsx)
            print(f"🧪 已生成样例: {sg_xlsx}")

    ensure_sample_excels()

    if au_xlsx.exists():
        choice = input("是否导入澳大利亚数据（AUSTRALIA.xlsx）？(y/n，默认y): ").strip().lower()
        if choice != 'n':
            clear_choice = input("是否清空AU现有数据？(y/n，默认n): ").strip().lower()
            import_au_from_excel(db, str(au_xlsx), clear_choice == 'y')
    if uk_xlsx.exists():
        choice = input("是否导入英国数据（UK.xlsx）？(y/n，默认y): ").strip().lower()
        if choice != 'n':
            clear_choice = input("是否清空UK现有数据？(y/n，默认n): ").strip().lower()
            import_uk_from_excel(db, str(uk_xlsx), clear_choice == 'y')
    if sg_xlsx.exists():
        choice = input("是否导入新加坡数据（SINGAPORE.xlsx）？(y/n，默认y): ").strip().lower()
        if choice != 'n':
            clear_choice = input("是否清空SG现有数据？(y/n，默认n): ").strip().lower()
            import_sg_from_excel(db, str(sg_xlsx), clear_choice == 'y')
    
    # 显示统计信息（仅美国数据集合）
    try:
    show_database_stats(db)
    except Exception:
        pass
    
    # 询问是否导出
    export_choice = input("\n是否导出当前数据到CSV？(y/n，默认n): ").strip().lower()
    if export_choice == 'y':
        export_to_csv(db)
    
    print("\n🎉 数据库管理完成！")

if __name__ == "__main__":
    main() 